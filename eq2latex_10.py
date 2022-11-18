# Première V d'une app pour convertir une cellule Excel convenablement mise
# en page (voir docu) en formule laTex imprimable. Selon le standard des
# rapports de laboratoire du cégep et de l'université.

import openpyxl
import sigfig
print(sigfig.round(1324, sigfigs = 3))

alphabet = ('A', 'B', 'C', 'D', 'E')  # À continuer


def strCell2xyCell(strCell):
    alphabet = ('A', 'B', 'C', 'D', 'E')  # À continuer
    x = ''
    y = 0
    l = len(strCell)
    for i in range(l):
        if strCell[i] in alphabet:
            x = x + strCell[i]
        else:
            chiffre = int(strCell[i])
            y = y+chiffre

    xyCell = [x, y]

    return xyCell


def cellShift(G_Ou_D, Cell):
    alphabet_list = ('A', 'B', 'C', 'D', 'E')  # À continuer
    if G_Ou_D == 'G':
        shiftvalue = -1
    elif G_Ou_D == 'D':
        shiftvalue = 1
    else:
        print('erreur cellshift')

    xycell = strCell2xyCell(Cell)

    index_lettre = alphabet_list.index(xycell[0])
    index_lettre = index_lettre + shiftvalue

    newcell = str(alphabet_list[index_lettre]) + str(xycell[1])


    #xycell[1] = xycell[1] + shiftvalue


    return newcell


def getdata(filepath, cell):
    wbData = openpyxl.load_workbook(filename = filepath,data_only=True)
    sh = wbData.active
    data = sh.value[cell]

    return data



filePath = input('Entrer le filepath du fichier xlsx : ')

wb = openpyxl.load_workbook(filePath)
sh = wb.active
print(sh)
mainCell = input('''Cellule à analyser (LETTREchiffre) jusqu'à ZZ99 : ''')
valueMainCell = sh[mainCell]
symboleMain = sh[cellShift('G', mainCell)]
unitMain = sh[cellShift('D', mainCell)]
datatest=getdata(filePath,mainCell)





print(mainCell, symboleMain, valueMainCell.value, unitMain,datatest)



"""
wb = openpyxl.load_workbook('test file.xlsx')
sh = wb.active

print(sh)

valueMainCell = sh['C2']

print(valueMainCell.value)"""
